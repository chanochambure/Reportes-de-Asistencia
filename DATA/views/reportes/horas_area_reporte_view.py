#!/usr/bin/env python
# -*- coding: utf-8 -*-

#Import de Librerias
import sys
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import datetime
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

#Import de Modulos
BASE_DIR='../../'
sys.path.insert(0,BASE_DIR)
from constants import *
from models import mark,lunchtime,trabajador
from controller import controller_lunchtime,controller_trabajador,controller_area,controller_tipo
from views.admin import control_marks_view
from views import control_single_mark_view
from reporte import reporte_horas_area

#str(a+datetime.timedelta(days=2))

class horas_area_reporte_view(QDialog):
	def __init__(self,parent=None):
		db=get_connection()
		self.lista_areas=[]
		self.lista_tipos=[]
		if(db==None):
			self.close()
		db.close()
		self.time_to_work=8*60
		self.reporte_matrix=[]
		super(horas_area_reporte_view, self).__init__(parent)
		self.horas_reporte_create()
		screenGeometry = QApplication.desktop().availableGeometry()
		self.resize(screenGeometry.width(), screenGeometry.height())
		self.showMaximized()
		self.setWindowTitle(REPORTES_HORAS_AREA_TITLE)
		self.show()

	def horas_reporte_create(self):
		#GRID
		grid = QGridLayout()

		#WIDGETS
		label_title = QLabel(REPORTES_HORAS_AREA_TITLE)
		label_date1 = QLabel(ADMIN_CONTROL_MARKS_DATE1)
		label_date2 = QLabel(ADMIN_CONTROL_MARKS_DATE2)
		label_area = QLabel(REPO_AREA_NAME_R_H_A)
		label_tipo = QLabel(REPO_TIPO_NAME_R_H_A)

		button_back = QPushButton(BUTTON_BACK,self)
		button_excel = QPushButton(BUTTON_EXCEL,self)
		button_reporte = QPushButton(BUTTON_REPORTE,self)

		self.a_box = QComboBox()
		self.t_box = QComboBox()

		self.d_box1 = QComboBox()
		self.m_box1 = QComboBox()
		self.y_box1 = QComboBox()
		self.d_box2 = QComboBox()
		self.m_box2 = QComboBox()
		self.y_box2 = QComboBox()

		self.reporte_table = QTableWidget()

		#Modificacion widgets
		font_title = QFont()
		font_title.setPointSize(FONT_TITLE_SIZE)
		label_title.setFont(font_title)

		self.create_combo_box()

		self.reporte_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
		self.clear_reporte_table()

		#FUNCTION
		self.connect(button_back, SIGNAL("clicked()"),self.close)
		self.connect(button_excel, SIGNAL("clicked()"),self.to_excel_function)
		self.connect(button_reporte, SIGNAL("clicked()"),self.generate_reporte)

		self.connect(self.y_box1, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box1)
		self.connect(self.m_box1, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box1)
		self.connect(self.y_box2, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box2)
		self.connect(self.m_box2, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box2)
		self.connect(self.a_box, SIGNAL("currentIndexChanged(QString)"), self.get_types_area)

		#GRID SIZE
		grid.setHorizontalSpacing(REPORTE_HORAS_AREA_X_GRID)
		grid.setVerticalSpacing(REPORTE_HORAS_AREA_Y_GRID)

		#WIDGETS TO GRID
		grid.addWidget(label_title,GRID_X_POSITION_TITLE_R_HORAS_AREA,GRID_Y_POSITION_TITLE_R_HORAS_AREA,1,3)
		grid.addWidget(button_back,GRID_X_POSITION_BUTTON_BACK_R_H_A,GRID_Y_POSITION_BUTTON_BACK_R_H_A)
		grid.addWidget(button_excel,GRID_X_POSITION_EXCEL_R_H_A,GRID_Y_POSITION_EXCEL_R_H_A)
		grid.addWidget(button_reporte,GRID_X_POSITION_REPO_R_H_A,GRID_Y_POSITION_REPO_R_H_A)
		grid.addWidget(label_date1,GRID_X_POSITION_DATEMARK_R_H_A,GRID_Y_POSITION_LABEL_DATE)
		grid.addWidget(label_date2,GRID_X_POSITION_DATEMARK_R_H_A+1,GRID_Y_POSITION_LABEL_DATE)
		grid.addWidget(label_area,GRID_X_POSITION_CONFIG_R_H_A,GRID_Y_POSITION_R_H_A_LABEL_A)
		grid.addWidget(label_tipo,GRID_X_POSITION_CONFIG_R_H_A,GRID_Y_POSITION_R_H_A_LABEL_T)

		grid.addWidget(self.a_box,GRID_X_POSITION_CONFIG_R_H_A,GRID_Y_POSITION_R_H_A_BOX_A)
		grid.addWidget(self.t_box,GRID_X_POSITION_CONFIG_R_H_A,GRID_Y_POSITION_R_H_A_BOX_T)

		grid.addWidget(self.d_box1,GRID_X_POSITION_DATEMARK_R_H_A,GRID_Y_POSITION_DAY_DATEMARK_R_H_A)
		grid.addWidget(self.m_box1,GRID_X_POSITION_DATEMARK_R_H_A,GRID_Y_POSITION_MONTH_DATEMARK_R_H_A)
		grid.addWidget(self.y_box1,GRID_X_POSITION_DATEMARK_R_H_A,GRID_Y_POSITION_YEAR_DATEMARK_R_H_A)
		grid.addWidget(self.d_box2,GRID_X_POSITION_DATEMARK_R_H_A+1,GRID_Y_POSITION_DAY_DATEMARK_R_H_A)
		grid.addWidget(self.m_box2,GRID_X_POSITION_DATEMARK_R_H_A+1,GRID_Y_POSITION_MONTH_DATEMARK_R_H_A)
		grid.addWidget(self.y_box2,GRID_X_POSITION_DATEMARK_R_H_A+1,GRID_Y_POSITION_YEAR_DATEMARK_R_H_A)

		grid.addWidget(self.reporte_table,GRID_X_POSITION_R_H_A_T,GRID_Y_POSITION_R_H_A_T,
						GRID_X_SIZE_R_H_A_T,GRID_Y_SIZE_R_H_A_T)

		#LAYAOUT
		self.setLayout(grid)

	def generate_reporte(self):
		day1=self.d_box1.currentText()
		month1=self.m_box1.currentText()
		year1=self.y_box1.currentText()
		day2=self.d_box2.currentText()
		month2=self.m_box2.currentText()
		year2=self.y_box2.currentText()
		datestr1=date_to_str(year1,month1,day1)
		datestr2=date_to_str(year2,month2,day2)
		ac_area=self.a_box.currentIndex()
		if(ac_area>0):
			ac_area=self.lista_areas[ac_area-1].id
		ac_tipo=self.t_box.currentIndex()
		if(ac_tipo>0):
			ac_tipo=self.lista_tipos[ac_tipo-1].id
		db=get_connection()
		if(db):
			self.reporte_matrix=reporte_horas_area.get_reporte_horas_area(db,ac_area,ac_tipo,datestr1,datestr2)
			self.insert_reporte_table(datestr1,datestr2)

	def to_excel_function(self):
		if(len(self.reporte_matrix)==0):
			QMessageBox.warning(self, 'Error',NONE_TO_SAVE, QMessageBox.Ok)
		else:
			self.to_excel()

	def create_combo_box(self):
		#Areas:
		db=get_connection()
		if(db):
			self.lista_areas=controller_area.get_r_areas(db)
			self.a_box.addItem("Todas")
			for i in self.lista_areas:
				self.a_box.addItem(i.name)
			self.t_box.addItem("Todas")
			db.close()
		#Dates:
		actual_date=datetime.datetime.now()
		year_list=range(actual_date.year - MORE_YEARS,actual_date.year + 1)
		year_list.reverse()
		for year in year_list:
			self.y_box1.addItem(str(year))
			self.y_box2.addItem(str(year))
		for month in range(1,13):
			str_month=str(month)
			if(month<10):
				str_month="0"+str_month
			self.m_box1.addItem(str_month)
			self.m_box2.addItem(str_month)
		self.m_box1.setCurrentIndex(actual_date.month - 1)
		self.m_box2.setCurrentIndex(actual_date.month - 1)
		self.max_day1=0
		self.max_day2=0
		if(actual_date.month==2):
			self.max_day1=int(actual_date.year%4==0)+28
			self.max_day2=int(actual_date.year%4==0)+28
		elif(actual_date.month==1 or actual_date.month==3 or actual_date.month==5 or actual_date.month==7 or actual_date.month==8 or actual_date.month==10 or actual_date.month==12):
			self.max_day1=31
			self.max_day2=31
		else:
			self.max_day1=30
			self.max_day2=30
		for day in range(1,self.max_day1+1):
			str_day=str(day)
			if(day<10):
				str_day="0"+str_day
			self.d_box1.addItem(str_day)
			self.d_box2.addItem(str_day)
		self.d_box1.setCurrentIndex(actual_date.day - 1)
		self.d_box2.setCurrentIndex(actual_date.day - 1)

	def day_combo_box1(self):
		current_day = int(self.d_box1.currentText())
		current_month = int(self.m_box1.currentText())
		current_year = int(self.y_box1.currentText())
		new_day=0
		if(current_month==2):
			new_day=int(current_year%4==0)+28
		elif(current_month==1 or current_month==3 or current_month==5 or current_month==7 or current_month==8 or current_month==10 or current_month==12):
			new_day=31
		else:
			new_day=30
		if(self.max_day1!=new_day):
			self.max_day1=new_day
			self.d_box1.clear()
			for day in range(1,self.max_day1+1):
				str_day=str(day)
				if(day<10):
					str_day="0"+str_day
				self.d_box1.addItem(str_day)
			if(current_day<=self.max_day1):
				self.d_box1.setCurrentIndex(current_day - 1)
			else:
				self.d_box1.setCurrentIndex(0)

	def day_combo_box2(self):
		current_day = int(self.d_box2.currentText())
		current_month = int(self.m_box2.currentText())
		current_year = int(self.y_box2.currentText())
		new_day=0
		if(current_month==2):
			new_day=int(current_year%4==0)+28
		elif(current_month==1 or current_month==3 or current_month==5 or current_month==7 or current_month==8 or current_month==10 or current_month==12):
			new_day=31
		else:
			new_day=30
		if(self.max_day2!=new_day):
			self.max_day2=new_day
			self.d_box2.clear()
			for day in range(1,self.max_day2+1):
				str_day=str(day)
				if(day<10):
					str_day="0"+str_day
				self.d_box2.addItem(str_day)
			if(current_day<=self.max_day2):
				self.d_box2.setCurrentIndex(current_day - 1)
			else:
				self.d_box2.setCurrentIndex(0)

	def get_types_area(self):
		current_id_area = self.a_box.currentIndex()
		self.t_box.clear()
		self.t_box.addItem("Todas")
		if(current_id_area>0):
			db=get_connection()
			if(db):
				self.lista_tipos=controller_tipo.get_r_tipos(db,self.lista_areas[current_id_area-1].id)
				for i in self.lista_tipos:
					self.t_box.addItem(i.name)
				db.close()

	def clear_reporte_table(self):
		self.reporte_table.clear();
		self.reporte_table.setRowCount(0)
		self.reporte_table.setColumnCount(0)

	def insert_reporte_table(self,d1,d2):
		self.clear_reporte_table()
		if(len(self.reporte_matrix)):
			i_date=to_date(d1,True)
			f_date=to_date(d2,True)
			self.fechas_lista=[]
			while(i_date<=f_date):
				self.fechas_lista.append(str(i_date))
				i_date=i_date + datetime.timedelta(days=1)
			if(self.fechas_lista):
				self.rows = len(self.reporte_matrix)
				self.reporte_table.setRowCount(self.rows)
				self.reporte_table.setColumnCount(SIZE_COLUMNS_TABLE_REPORTE_HORAS_AREA+len(self.fechas_lista))
				self.reporte_table.setHorizontalHeaderLabels(REPORTE_HORAS_AREA_HEADER+self.fechas_lista)
				stringVert = []
				for index_report in range(len(self.reporte_matrix)):
					self.reporte_table.setItem(index_report,0, QTableWidgetItem(self.reporte_matrix[index_report][0]))
					self.reporte_table.setItem(index_report,1, QTableWidgetItem(self.reporte_matrix[index_report][1]))
					for fecha_index in range(len(self.fechas_lista)):
						value_f=VALUE_NOT_WORK
						if(self.fechas_lista[fecha_index] in self.reporte_matrix[index_report][2]):
							value_f= self.reporte_matrix[index_report][2][self.fechas_lista[fecha_index]]
						self.reporte_table.setItem(index_report,fecha_index+2, QTableWidgetItem(value_f))
				self.reporte_table.setVerticalHeaderLabels(stringVert)
				self.reporte_table.resizeColumnsToContents()

	def to_excel(self):
		dest_filename = unicode(QFileDialog.getSaveFileName(self, SAVE_FILE_TITLE, '', ".xlsx(*.xlsx)"))
		wb = Workbook()
		ws1 = wb.active
		ws1.title=REPORTES_HORAS_AREA_TITLE
		ws1["B1"]=REPORTES_HORAS_AREA_TITLE
		ws1["A3"]=REPORTE_HORAS_AREA_HEADER[0]
		ws1["B3"]=REPORTE_HORAS_AREA_HEADER[1]
		for index_f in range(len(self.fechas_lista)):
			try:
				celda = get_column_letter(index_f+3)+str(3)
				ws1[celda] = self.fechas_lista[index_f]
			except AttributeError:
				pass
		for currentRow in range(self.reporte_table.rowCount()):
			for currentColumn in range(self.reporte_table.columnCount()):
				try:
					teext = str(self.reporte_table.item(currentRow,currentColumn).text())
					celda = get_column_letter(currentColumn+1)+str(currentRow+4)
					ws1[celda] = teext
				except AttributeError:
					pass
		try:
			wb.save(filename = dest_filename)
			QMessageBox.question(self, 'Message',CREATE_EXCEL_SUCCESS,QMessageBox.Ok)
		except IOError:
			QMessageBox.warning(self, 'Error',EXCEL_PROBLEM, QMessageBox.Ok)