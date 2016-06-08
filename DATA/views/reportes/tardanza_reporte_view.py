#!/usr/bin/env python
# -*- coding: utf-8 -*-

#Import de Librerias
import sys
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import datetime
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

#Import de Modulos
BASE_DIR='../../'
sys.path.insert(0,BASE_DIR)
from constants import *
from models import mark,lunchtime,trabajador
from controller import controller_lunchtime,controller_trabajador
from views.admin import control_marks_view
from views import control_single_mark_view
from reporte import reporte_tardanza

class tardanza_reporte_view(QDialog):
	def __init__(self,worker_to_set,parent=None):
		self.time_to_work=8*60
		self.var_singleton=False
		self.ventana=None
		self.reporte_matrix=[]
		super(tardanza_reporte_view, self).__init__(parent)
		self.worker=worker_to_set
		self.tardanza_reporte_create()
		screenGeometry = QApplication.desktop().availableGeometry()
		self.resize(screenGeometry.width(), screenGeometry.height())
		self.showMaximized()
		self.setWindowTitle(REPORTES_TARDANZA_TITLE)
		self.show()

	def tardanza_reporte_create(self):
		#GRID
		grid = QGridLayout()

		#WIDGETS
		label_title = QLabel(REPORTES_TARDANZA_TITLE)
		label_date1 = QLabel(ADMIN_CONTROL_MARKS_DATE1)
		label_date2 = QLabel(ADMIN_CONTROL_MARKS_DATE2)
		label_name = QLabel(REPORTE_LABEL_NAME)
		label_total_horas = QLabel(REPORTE_LABEL_TOTAL_MINUTOS_TARDE)

		self.text_name = QLineEdit()
		self.text_total_horas = QLineEdit()

		button_reporte = QPushButton(BUTTON_REPORTE,self)
		button_control_mark = QPushButton(BUTTON_CONTROL_MARK,self)
		button_excel = QPushButton(BUTTON_EXCEL,self)
		button_back = QPushButton(BUTTON_BACK,self)

		self.d_box1 = QComboBox()
		self.m_box1 = QComboBox()
		self.y_box1 = QComboBox()
		self.d_box2 = QComboBox()
		self.m_box2 = QComboBox()
		self.y_box2 = QComboBox()

		self.reporte_table = QTableWidget()

		#Modificacion widgets
		font_title = QFont()
		font_title.setPointSize(FONT_TITLE_SIZE)
		label_title.setFont(font_title)

		self.text_name.setText(self.worker.pin+" - "+self.worker.name+" "+self.worker.father_last_name+" "+self.worker.mother_last_name)
		self.text_name.setReadOnly(True)
		self.text_total_horas.setText(REPORTE_TOTAL_HORAS_EMPTY)
		self.text_total_horas.setReadOnly(True)
		self.text_total_horas.setFixedSize(REPORTES_HORAS_BUTTON_SIZE_X,button_reporte.height())

		button_control_mark.setFixedSize(REPORTES_HORAS_BUTTON_SIZE_X,REPORTES_HORAS_BUTTON_SIZE_Y)
		button_excel.setFixedSize(REPORTES_HORAS_BUTTON_SIZE_X,REPORTES_HORAS_BUTTON_SIZE_Y)
		button_back.setFixedSize(REPORTES_HORAS_BUTTON_SIZE_X,REPORTES_HORAS_BUTTON_SIZE_Y)

		self.create_combo_box()

		self.reporte_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
		self.clear_reporte_table()
		header = self.reporte_table.horizontalHeader()
		header.setResizeMode(QHeaderView.Stretch)
		self.reporte_table.setVerticalHeaderLabels([''])

		#FUNCTION
		self.connect(button_back, SIGNAL("clicked()"),self.close)
		self.connect(button_reporte, SIGNAL("clicked()"),self.generate_reporte)
		self.connect(button_control_mark, SIGNAL("clicked()"),self.open_control_marks)
		self.connect(button_excel, SIGNAL("clicked()"),self.to_excel_function)
		self.connect(self.y_box1, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box1)
		self.connect(self.m_box1, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box1)
		self.connect(self.y_box2, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box2)
		self.connect(self.m_box2, SIGNAL("currentIndexChanged(QString)"), self.day_combo_box2)

		#GRID SIZE
		grid.setHorizontalSpacing(REPORTE_HORAS_X_GRID)
		grid.setVerticalSpacing(REPORTE_HORAS_Y_GRID)

		#WIDGETS TO GRID
		grid.addWidget(label_title,GRID_X_POSITION_TITLE_REP_HORAS,GRID_Y_POSITION_TITLE_REP_HORAS)
		grid.addWidget(label_name,GRID_X_POSITION_NAME_REP_HORAS,GRID_Y_POSITION_NAME_REP_HORAS)
		grid.addWidget(label_total_horas,GRID_X_POSITION_TOTAL_LABEL_REP_HORAS,GRID_Y_POSITION_TOTAL_LABEL_REP_HORAS)
		grid.addWidget(label_date1,GRID_X_POSITION_DATEMARK_REP_HORAS,GRID_Y_POSITION_LABEL_DATE)
		grid.addWidget(label_date2,GRID_X_POSITION_DATEMARK_REP_HORAS+1,GRID_Y_POSITION_LABEL_DATE)

		grid.addWidget(button_back,GRID_X_POSITION_BACK_REP_HORAS,GRID_Y_POSITION_BACK_REP_HORAS)
		grid.addWidget(button_reporte,GRID_X_POSITION_REPO_REP_HORAS,GRID_Y_POSITION_REPO_REP_HORAS)
		grid.addWidget(button_excel,GRID_X_POSITION_EXCEL_REP_HORAS,GRID_Y_POSITION_EXCEL_REP_HORAS)
		grid.addWidget(button_control_mark,GRID_X_POSITION_CMARK_REP_HORAS,GRID_Y_POSITION_CMARK_REP_HORAS)

		grid.addWidget(self.text_name,GRID_X_POSITION_NAME_REP_HORAS,GRID_Y_POSITION_TEXT_N_REP_HORAS,1,GRID_Y_POSITION_TEXT_N_REP_HORAS+2)
		grid.addWidget(self.text_total_horas,GRID_X_POSITION_TOTAL_REP_HORAS,GRID_Y_POSITION_TOTAL_REP_HORAS)

		grid.addWidget(self.d_box1,GRID_X_POSITION_DATEMARK_REP_HORAS,GRID_Y_POSITION_DAY_DATEMARK_REP_HORAS)
		grid.addWidget(self.m_box1,GRID_X_POSITION_DATEMARK_REP_HORAS,GRID_Y_POSITION_MONTH_DATEMARK_REP_HORAS)
		grid.addWidget(self.y_box1,GRID_X_POSITION_DATEMARK_REP_HORAS,GRID_Y_POSITION_YEAR_DATEMARK_REP_HORAS)
		grid.addWidget(self.d_box2,GRID_X_POSITION_DATEMARK_REP_HORAS+1,GRID_Y_POSITION_DAY_DATEMARK_REP_HORAS)
		grid.addWidget(self.m_box2,GRID_X_POSITION_DATEMARK_REP_HORAS+1,GRID_Y_POSITION_MONTH_DATEMARK_REP_HORAS)
		grid.addWidget(self.y_box2,GRID_X_POSITION_DATEMARK_REP_HORAS+1,GRID_Y_POSITION_YEAR_DATEMARK_REP_HORAS)

		grid.addWidget(self.reporte_table,GRID_X_POSITION_REPORTE_HORAS_TABLE_1,GRID_Y_POSITION_REPORTE_HORAS_TABLE_1,
						GRID_X_POSITION_REPORTE_HORAS_TABLE_2,GRID_Y_POSITION_REPORTE_HORAS_TABLE_2)

		#LAYAOUT
		self.setLayout(grid)

	def create_combo_box(self):
		actual_date=datetime.datetime.now()
		year_list=range(actual_date.year - MORE_YEARS,actual_date.year + 1)
		year_list.reverse()
		for year in year_list:
			self.y_box1.addItem(str(year))
			self.y_box2.addItem(str(year))
		for month in range(1,13):
			str_month=str(month)
			if(month<10):
				str_month="0"+str_month
			self.m_box1.addItem(str_month)
			self.m_box2.addItem(str_month)
		self.m_box1.setCurrentIndex(actual_date.month - 1)
		self.m_box2.setCurrentIndex(actual_date.month - 1)
		self.max_day1=0
		self.max_day2=0
		if(actual_date.month==2):
			self.max_day1=int(actual_date.year%4==0)+28
			self.max_day2=int(actual_date.year%4==0)+28
		elif(actual_date.month==1 or actual_date.month==3 or actual_date.month==5 or actual_date.month==7 or actual_date.month==8 or actual_date.month==10 or actual_date.month==12):
			self.max_day1=31
			self.max_day2=31
		else:
			self.max_day1=30
			self.max_day2=30
		for day in range(1,self.max_day1+1):
			str_day=str(day)
			if(day<10):
				str_day="0"+str_day
			self.d_box1.addItem(str_day)
			self.d_box2.addItem(str_day)
		self.d_box1.setCurrentIndex(actual_date.day - 1)
		self.d_box2.setCurrentIndex(actual_date.day - 1)

	def day_combo_box1(self):
		current_day = int(self.d_box1.currentText())
		current_month = int(self.m_box1.currentText())
		current_year = int(self.y_box1.currentText())
		new_day=0
		if(current_month==2):
			new_day=int(current_year%4==0)+28
		elif(current_month==1 or current_month==3 or current_month==5 or current_month==7 or current_month==8 or current_month==10 or current_month==12):
			new_day=31
		else:
			new_day=30
		if(self.max_day1!=new_day):
			self.max_day1=new_day
			self.d_box1.clear()
			for day in range(1,self.max_day1+1):
				str_day=str(day)
				if(day<10):
					str_day="0"+str_day
				self.d_box1.addItem(str_day)
			if(current_day<=self.max_day1):
				self.d_box1.setCurrentIndex(current_day - 1)
			else:
				self.d_box1.setCurrentIndex(0)

	def day_combo_box2(self):
		current_day = int(self.d_box2.currentText())
		current_month = int(self.m_box2.currentText())
		current_year = int(self.y_box2.currentText())
		new_day=0
		if(current_month==2):
			new_day=int(current_year%4==0)+28
		elif(current_month==1 or current_month==3 or current_month==5 or current_month==7 or current_month==8 or current_month==10 or current_month==12):
			new_day=31
		else:
			new_day=30
		if(self.max_day2!=new_day):
			self.max_day2=new_day
			self.d_box2.clear()
			for day in range(1,self.max_day2+1):
				str_day=str(day)
				if(day<10):
					str_day="0"+str_day
				self.d_box2.addItem(str_day)
			if(current_day<=self.max_day2):
				self.d_box2.setCurrentIndex(current_day - 1)
			else:
				self.d_box2.setCurrentIndex(0)

	def closeEvent(self, evnt):
		if(self.ventana):
			self.ventana.close()

	def generate_reporte(self):
		if(self.var_singleton):
			QMessageBox.warning(self, 'Error',CONTROL_MARK_OPENED, QMessageBox.Ok)
		else:
			day1=self.d_box1.currentText()
			month1=self.m_box1.currentText()
			year1=self.y_box1.currentText()
			day2=self.d_box2.currentText()
			month2=self.m_box2.currentText()
			year2=self.y_box2.currentText()
			datestr1=date_to_str(year1,month1,day1)
			datestr2=date_to_str(year2,month2,day2)
			db=get_connection()
			if(db):
				self.reporte_matrix=reporte_tardanza.get_reporte_tardanza(self.worker.pin,datestr1,datestr2,db)
				self.minutos=reporte_tardanza.get_total_minutes(self.reporte_matrix)
				self.insert_reporte_table()

	def to_excel_function(self):
		if(self.var_singleton):
			QMessageBox.warning(self, 'Error',CONTROL_MARK_OPENED, QMessageBox.Ok)
		elif(len(self.reporte_matrix)==0):
			QMessageBox.warning(self, 'Error',NONE_TO_SAVE, QMessageBox.Ok)
		else:
			self.var_singleton=True
			self.to_excel()
			self.ventana=None
			self.var_singleton=False

	def to_excel(self):
		dest_filename = unicode(QFileDialog.getSaveFileName(self, SAVE_FILE_TITLE, '', ".xlsx(*.xlsx)"))
		wb = Workbook()
		ws1 = wb.active
		ws1.title=REPORTES_TARDANZA_TITLE
		ws1["B1"]=REPORTES_TARDANZA_TITLE
		ws1["A2"]=REPORTE_LABEL_NAME
		ws1["B2"]=self.worker.name+" "+self.worker.father_last_name+" "+self.worker.mother_last_name
		ws1["C2"]=self.worker.pin
		for currentColumn in range(self.reporte_table.columnCount()):
			try:
				celda = get_column_letter(currentColumn+1)+str(3)
				ws1[celda] = REPORTE_TARDANZA_TITLE_ROWS[currentColumn]
			except AttributeError:
				pass
		counter=4
		for currentRow in range(self.reporte_table.rowCount()):
			for currentColumn in range(self.reporte_table.columnCount()-1):
				try:
					teext = str(self.reporte_table.item(currentRow,currentColumn).text())
					celda = get_column_letter(currentColumn+1)+str(currentRow+4)
					ws1[celda] = teext
				except AttributeError:
					pass
			counter+=1
		celda = "A"+str(counter)
		ws1[celda] = REPORTE_LABEL_TOTAL_MINUTOS_TARDE
		celda = "B"+str(counter)
		ws1[celda] = self.text_total_horas.text()
		try:
			wb.save(filename = dest_filename)
			QMessageBox.question(self, 'Message',CREATE_EXCEL_SUCCESS,QMessageBox.Ok)
		except IOError:
			QMessageBox.warning(self, 'Error',EXCEL_PROBLEM, QMessageBox.Ok)

	def open_control_marks(self):
		if(self.var_singleton):
			QMessageBox.warning(self, 'Error',CONTROL_MARK_OPENED, QMessageBox.Ok)
		else:
			self.var_singleton=True
			self.ventana=control_marks_view.control_marks_view(self.worker)
			self.ventana.exec_()
			self.ventana=None
			self.var_singleton=False

	def insert_reporte_table(self):
		self.clear_reporte_table()
		if(self.minutos!=None):
			self.text_total_horas.setText(self.minutos)
		else:
			self.text_total_horas.setText(REPORTE_TOTAL_HORAS_EMPTY)
		if(len(self.reporte_matrix)):
			self.rows = len(self.reporte_matrix)
			self.reporte_table.setRowCount(self.rows)
			stringVert = []
			for index_report in range(len(self.reporte_matrix)):
				self.reporte_table.setItem(index_report,0, QTableWidgetItem(self.reporte_matrix[index_report][0]))
				self.reporte_table.setItem(index_report,1, QTableWidgetItem(str(self.reporte_matrix[index_report][1])))
				self.reporte_table.setItem(index_report,2, QTableWidgetItem(mins_to_str_time_ot(self.reporte_matrix[index_report][2])))
				self.reporte_table.setItem(index_report,3, QTableWidgetItem(mins_to_str_time_ot(self.reporte_matrix[index_report][3])))
				self.reporte_table.setItem(index_report,4, QTableWidgetItem(mins_to_str_time_ot(self.reporte_matrix[index_report][4])))
				self.reporte_table.setItem(index_report,5, QTableWidgetItem(mins_to_str_time_ot(self.reporte_matrix[index_report][5])))
				self.btn_sell = QPushButton(MODIFICAR_REPORTE_MARKS)
				self.btn_sell.clicked.connect(self.selection)
				self.reporte_table.setCellWidget(index_report,6,self.btn_sell)
				stringVert.append(str(index_report+1))
			self.reporte_table.setVerticalHeaderLabels(stringVert)

	def clear_reporte_table(self):
		self.reporte_table.clear();
		self.reporte_table.setRowCount(0);
		self.reporte_table.setColumnCount(SIZE_COLUMNS_TABLE_REPORTE_TARDANZA)
		self.reporte_table.setHorizontalHeaderLabels(REPORTE_TARDANZA_TITLE_ROWS)

	def selection(self):
		if(self.var_singleton):
			QMessageBox.warning(self, 'Error',CONTROL_MARK_OPENED, QMessageBox.Ok)
		else:
			index = self.reporte_table.indexAt(qApp.focusWidget().pos())
			if index.isValid():
				date = str(self.reporte_table.item(index.row(),0).text())
				self.var_singleton=True
				self.ventana=control_single_mark_view.control_single_mark_view(self.worker,date)
				self.ventana.exec_()
				self.ventana=None
				self.var_singleton=False